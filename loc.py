import time

import keyboard

import pandas as pd
import pyperclip as pc
import openpyxl
import unicodedata

# Selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# PyautoGUI
import pyautogui as pg


def parar_execucao():
    """
        Função para parar a execução do script quando manter pressionado a tecla X
    """
    return keyboard.is_pressed('x')


# Start Configuração
espera = True
cont = 1
linha = 2
conseguiu_localizar_ponto=False
tempo_confidence_padrao = 0.9
confidence_image = tempo_confidence_padrao
num_tentativas = 1

STR_STATUS = '[ STATUS ]'
STR_RESULTADO = '[ RESULTADO ]'
NOME_PLANILHA = r"gelocalizacao_camed.xlsx"
URL_MAPS = 'https://www.google.com/maps/'
TEMPO_SLEEP_MIN = 1
TEMPO_SLEEP_MAX = 3

options = Options()
options.add_experimental_option("detach", True)

service = Service(ChromeDriverManager().install())
# End Configuração


# Start
# Abre a planilha
planilha = openpyxl.load_workbook(NOME_PLANILHA)
colunas = planilha['Planilha1']

# Abre o Maps
browser = webdriver.Chrome(service=service, options=options)
browser.get(URL_MAPS)
# browser.fullscreen_window()
time.sleep(5)

# Start Loop
for coluna in colunas:
    # Verifica para parar o script
    if parar_execucao():
        browser.close()
        break

    # Ativa a planilha
    ws = planilha.active
    celula_endereco = ws['B'+str(linha)].value
    # time.sleep(TEMPO_SLEEP_MIN)

    # Pesquisa o endereço
    print(cont, STR_STATUS, 'Pesquisando endereço')
    while espera:
        try:
            pesquisar =  browser.find_element(By.XPATH, '//*[@id="searchboxinput"]')
            pesquisar.click()
            pesquisar.send_keys(
                celula_endereco,
                Keys.ENTER
            )
            espera = False
        except:
            time.sleep(TEMPO_SLEEP_MAX)

    espera=True
    
    # Rigth Click na localização
    print(cont, STR_STATUS, 'Rigth Click na localização')
    while espera:
        try:
            time.sleep(TEMPO_SLEEP_MAX)
            print(cont, STR_STATUS, 'Tentativa:', num_tentativas)
            ponto = pg.locateCenterOnScreen('ponto.png', confidence=confidence_image)
            pg.rightClick(ponto.x, ponto.y)
            conseguiu_localizar_ponto=True
            espera = False
        except:
            if (num_tentativas <= 2):
                # confidence_image -= 0.1
                num_tentativas+=1
            else:
                espera = False
                break

    num_tentativas=1
    confidence_image=tempo_confidence_padrao
    espera=True

    if (conseguiu_localizar_ponto):
        # Copia a Latitude e Longitude
        print(cont, STR_STATUS, 'Copia a Latitude e Longitude')
        while espera:
            try:
                time.sleep(TEMPO_SLEEP_MIN)
                print(cont, STR_STATUS, 'Tentativa:', num_tentativas)
                # time.sleep(TEMPO_SLEEP_MIN)
                latitude_longitude_click =  browser.find_element(By.XPATH,
                                                                 '//*[@id="action-menu"]/div[1]')
                # Copia
                latitude_longitude_click.click()
                # Cola
                latitude_longitude = pc.paste()
                espera = False
            except:
                num_tentativas+=1

        num_tentativas=1
        espera=True
    else:
        latitude_longitude = 'VAZIO'
        print(cont, STR_STATUS, 'Não foi possivel localizar o ponto da localização')

    # Mostra execucao no Console
    print(cont, STR_RESULTADO, latitude_longitude)
  
    # Salvando Latitude e Longitude na planilha
    print(cont, STR_STATUS, 'Salvando planilha')
    ws['C'+str(linha)].value = latitude_longitude
    # time.sleep(TEMPO_SLEEP_MAX)
    planilha.save(NOME_PLANILHA)
        
    # Limpando o pesquisar
    print(cont, STR_STATUS, 'Limpando o pesquisar\n')
    while espera:
        try:
            pesquisar.clear()
            espera = False
        except:
            time.sleep(TEMPO_SLEEP_MAX)

    espera=True
    linha += 1
    cont += 1
    conseguiu_localizar_ponto=False
    # time.sleep(5)
    # End Loop










# Abre Google Maps
# browser.get(URL_MAPS)
# time.sleep(2)

# Pesquisar
# while espera:
#     try:
#         pesquisar =  browser.find_element(By.XPATH, '//*[@id="searchboxinput"]')
#         pesquisar.click()
#         pesquisar.send_keys(
#             "RUA HUMBERTO DE CAMPOS, GRACA, 11, SALVADOR, BAHIA, BA",
#             Keys.ENTER
#         )
#         break
#     except:
#         print('Não encontrei o campo de pesquisa')

# Localização
# while espera:
#     try:
#         ponto = pg.locateCenterOnScreen('ponto.png', confidence=0.7)
#         pg.rightClick(ponto.x, ponto.y)
#         break
#     except:
#         print('Não encontrei o ponto de localização')

# Latitude e Longitude
# while espera:
#     try:
#         latitude_longitude =  browser.find_element(By.XPATH, '//*[@id="action-menu"]/div[1]')
#         latitude_longitude.click()
#         break
#     except:
#         print('Não encontrei o campo da latitude e logintude')


# X
# while espera:
#     try:
#         x = browser.find_element(By.XPATH, '//*[@id="searchbox"]/div[3]/button')
#         x.click()
#         break
#     except:
#         time.sleep(1)
#         print('Não encontrei o campo do X')
